"""课表 Excel/CSV 解析。

支持两种布局：
1) 教务系统格式：单元格 = "课程名/(节次)周次/ 地点/ 教师/ 教学班;组成"
   —— 按 "/" 拆字段，节次在单元格里（如 (3-4节)1-16周），教学班用 ";" 分隔。
2) 经典「节次 × 星期」格式：单元格 = "课程名\\n教室\\n老师"，时间按行推断。

解析原则：课程名永远优先取第一个有效字段；解析不到/存疑的情况进 warnings，
绝不静默回退成整格原文或默认周次。
"""
from __future__ import annotations

import csv
import io
import re

WEEKDAY_NAMES = {
    "周一": 1, "星期一": 1, "礼拜一": 1,
    "周二": 2, "星期二": 2, "礼拜二": 2,
    "周三": 3, "星期三": 3, "礼拜三": 3,
    "周四": 4, "星期四": 4, "礼拜四": 4,
    "周五": 5, "星期五": 5, "礼拜五": 5,
    "周六": 6, "星期六": 6, "礼拜六": 6,
    "周日": 7, "周天": 7, "星期日": 7, "星期天": 7, "礼拜日": 7, "礼拜天": 7,
    "一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6, "日": 7,
    "mon": 1, "monday": 1, "tue": 2, "tuesday": 2, "wed": 3, "wednesday": 3,
    "thu": 4, "thursday": 4, "fri": 5, "friday": 5, "sat": 6, "saturday": 6,
    "sun": 7, "sunday": 7,
}

# 默认节次（经典格式解析不到时间时按行序用）
DEFAULT_SLOTS = [
    ("08:00", "09:40"), ("10:00", "11:40"), ("14:00", "15:40"),
    ("16:00", "17:40"), ("19:00", "20:40"),
]

DASH = r"[-~～－—至]"
TIME_RE = re.compile(r"(\d{1,2}[:：]\d{2})\s*" + DASH + r"\s*(\d{1,2}[:：]\d{2})")
WEEKS_RE = re.compile(r"(?:第)?\s*(\d{1,2})\s*" + DASH + r"\s*(\d{1,2})\s*周")
SINGLE_WEEK_RE = re.compile(r"(?:第)?\s*(\d{1,2})\s*周")
MULTI_WEEKS_RE = re.compile(r"\d{1,2}\s*" + DASH + r"\s*\d{1,2}\s*周")
ODD_EVEN_RE = re.compile(r"单双周")            # 语义：每周都上
ODD_RE = re.compile(r"单周|\(单\)|（单）")
EVEN_RE = re.compile(r"双周|\(双\)|（双）")
WEEK_STRIP_RE = re.compile(
    r"[（(]?\s*(?:第)?\s*\d{1,2}\s*" + DASH + r"\s*\d{1,2}\s*周\s*[）)]?|"
    r"单双周|单周|双周|\(单\)|（单）|\(双\)|（双）|"
    r"(?:第)?\s*\d{1,2}\s*周"
)
SESSION_RE = re.compile(r"[（(]?\s*\d{1,2}\s*" + DASH + r"\s*\d{1,2}\s*节\s*[）)]?")
EMPTY_PAREN_RE = re.compile(r"[（(][）)]")
# 注意：不含"教"字——会误伤"王教练/教授"这类教师名；教室/教学楼已有"室/楼"覆盖
LOCATION_RE = re.compile(r"(楼|馆|室|中心|厂|讲堂|实验室|基地|区|栋|阶|堂|厅|号|群)")
# 班级名：如 动力2503、动力2502班（4 位数字；地点如"三山楼505"优先于班级判断）
CLASS_RE = re.compile(r"^[\u4e00-\u9fa5]{1,4}\d{4}(班)?$")

# 节次 → 开始时间（常见校历约定，可在导入后手动微调）
SESSION_START = {
    1: "08:00", 2: "08:55", 3: "10:10", 4: "11:05",
    5: "14:00", 6: "14:55", 7: "16:10", 8: "17:05",
    9: "19:00", 10: "19:55", 11: "20:50",
}
# 经典格式：上午/下午/晚上 × 第几大节 → 节次
PERIOD_SESSION = {
    ("上午", "一"): (1, 2), ("上午", "二"): (3, 4),
    ("下午", "三"): (5, 6), ("下午", "四"): (7, 8),
    ("晚上", "五"): (9, 11), ("晚上", "六"): (9, 11),
}

PALETTE = ["#e8893c", "#4a7fc7", "#4c9e76", "#c25b5b", "#a884d6", "#c98a2d", "#cf7a9d", "#4fa8a0"]


def parse_file(filename: str, data: bytes) -> dict:
    ext = (filename or "").lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext == "xlsx":
        return _parse_xlsx(data)
    if ext == "xls":
        return _parse_xls(data)
    if ext == "csv":
        return _parse_csv(data)
    raise ValueError("仅支持 .xls / .xlsx / .csv 文件")


def _normalize(v) -> str:
    if v is None:
        return ""
    return str(v).strip().replace("\u3000", " ")


def _find_header(rows):
    """找表头行：包含 >=2 个星期名的行，返回 (行号, {列号: 星期数})。"""
    for i, row in enumerate(rows[:15]):
        mapping = {}
        for ci, cell in enumerate(row):
            t = _normalize(cell).lower()
            for name, wd in WEEKDAY_NAMES.items():
                if len(name) == 1:
                    if t == name:
                        mapping[ci] = wd
                        break
                elif t == name or t.startswith(name):
                    mapping[ci] = wd
                    break
        if len(mapping) >= 3:
            return i, mapping
    return None, None


def _time_add(t: str, minutes: int) -> str:
    h, m = t.split(":")
    total = int(h) * 60 + int(m) + minutes
    return f"{total // 60:02d}:{total % 60:02d}"


def _pad_time(t: str) -> str:
    try:
        h, m = t.replace("：", ":").split(":")
        return f"{int(h):02d}:{int(m):02d}"
    except ValueError:
        return t


def _session_time(session):
    if not session:
        return None
    start, end = session
    s = SESSION_START.get(start)
    e = SESSION_START.get(end)
    if not s or not e:
        return None
    return s, _time_add(e, 45)


def _extract_weeks(text: str, warnings: list):
    """返回 (week_start, week_end, week_type)；解析不到时返回 None 并告警。"""
    week_type = "every"
    if ODD_EVEN_RE.search(text):
        week_type = "every"  # "单双周"= 每周都上
    elif ODD_RE.search(text):
        week_type = "odd"
    elif EVEN_RE.search(text):
        week_type = "even"
    m = WEEKS_RE.search(text)
    if m:
        return int(m.group(1)), int(m.group(2)), week_type
    sm = SINGLE_WEEK_RE.search(text)
    if sm:
        n = int(sm.group(1))
        return n, n, week_type
    if "周" in text:
        warnings.append(f"未能识别周次格式「{text.strip()[:30]}」，默认 1-20 周，请导入后核对")
    return None


def _parse_course_cell(cell_text: str, weekday: int, fallback_time, color: str, warnings: list) -> list:
    """一个格子可能含多门课（如 \\r\\n 分隔的工程训练冷/热），返回课程列表。"""
    if not cell_text:
        return []
    lines = [ln for ln in re.split(r"\r?\n", cell_text) if ln.strip()]
    out = []
    for line in lines:
        c = _parse_one_course(line.strip(), weekday, fallback_time, color, warnings)
        if c:
            out.append(c)
    return out


def _parse_one_course(text: str, weekday: int, fallback_time, color: str, warnings: list) -> dict | None:
    if not text or text in ("无", "-", "/", "／"):
        return None

    # 周次 & 节次（可能嵌在单元格任意位置）
    week_start, week_end, week_type = 1, 20, "every"
    w = _extract_weeks(text, warnings)
    if w:
        week_start, week_end, week_type = w
    if len(MULTI_WEEKS_RE.findall(text)) > 1:
        warnings.append(f"「{text.strip()[:30]}」含多段周次（如 1-8,10-16 周），只取了首段，请核对")
    session = None
    sm = re.search(r"(\d{1,2})\s*" + DASH + r"\s*(\d{1,2})节", text)
    if sm:
        session = (int(sm.group(1)), int(sm.group(2)))
    start_time, end_time = _session_time(session) or fallback_time or ("08:00", "09:40")

    name = teacher = location = ""

    if "/" in text:
        # —— 教务系统格式：课程/周次/地点/教师/教学班 ——
        for f in text.split("/"):
            f = f.strip()
            if not f:
                continue
            if ";" in f:
                continue  # 教学班组成，跳过
            clean = WEEK_STRIP_RE.sub("", f).strip("，,、;； ")
            clean = SESSION_RE.sub("", clean).strip("，,、;； ")
            clean = EMPTY_PAREN_RE.sub("", clean).strip()
            if not clean:
                continue
            if not name:
                name = clean  # 课程名优先：即使含数字/地点词
            elif LOCATION_RE.search(clean):
                location = location + ("·" if location else "") + clean
            elif CLASS_RE.match(clean):
                continue  # 班级名（4 位数字，如 动力2503）
            elif re.search(r"\d", clean):
                location = location + ("·" if location else "") + clean  # 含数字 → 地点/设施
            else:
                teacher = teacher + ("、" if teacher else "") + clean.replace(",", "、").replace("，", "、")
        if not name:
            name = WEEK_STRIP_RE.sub("", text).strip()
            warnings.append(f"课程名疑似解析异常「{text.strip()[:30]}」")
    else:
        # —— 经典格式：课程名 / 教室 / 老师 ——
        segs = [s for s in re.split(r"[\n;；]", text) if s.strip()]
        for s in segs:
            s2 = WEEK_STRIP_RE.sub("", s).strip("，,、;； ")
            if not s2:
                continue
            if not name:
                name = s2
            elif LOCATION_RE.search(s2) or re.search(r"\d{2,}", s2):
                location = location + ("·" if location else "") + s2
            else:
                teacher = teacher + ("、" if teacher else "") + s2.replace(",", "、").replace("，", "、")

    if not name:
        return None
    return {
        "name": name, "teacher": teacher, "location": location,
        "weekday": weekday, "start_time": start_time, "end_time": end_time,
        "week_start": week_start, "week_end": week_end, "week_type": week_type,
        "color": color,
    }


def _rows_to_courses(rows, header_idx, col_map):
    courses = []
    warnings: list = []
    used = set()
    time_idx = 0
    palette_i = 0
    for r in range(header_idx + 1, len(rows)):
        row = rows[r] if r < len(rows) else []
        if not row or not any(str(c) for c in row):
            continue  # 空行跳过（不推进节次计数）

        first_text = _normalize(row[0]) if row else ""
        t = _extract_time(first_text)
        slot = t or DEFAULT_SLOTS[time_idx % len(DEFAULT_SLOTS)]
        time_idx += 1
        period = first_text
        idx = _normalize(row[1]) if len(row) > 1 else ""
        fallback_time = _session_time(PERIOD_SESSION.get((period, idx))) or slot

        for ci, wd in col_map.items():
            if ci >= len(row):
                continue
            cell = row[ci]
            if cell is None:
                continue
            parsed = _parse_course_cell(str(cell), wd, fallback_time, PALETTE[palette_i % len(PALETTE)], warnings)
            for course in parsed:
                key = (course["name"], wd, course["start_time"],
                       course["week_start"], course["week_end"], course["week_type"])
                if key in used:
                    continue
                used.add(key)
                palette_i += 1
                courses.append(course)
    return courses, warnings


def _extract_time(text: str):
    m = TIME_RE.search(text)
    if m:
        return _pad_time(m.group(1)), _pad_time(m.group(2))
    return None


def _parse_xlsx(data: bytes) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.active
    rows = [[_normalize(c.value) for c in row] for row in ws.iter_rows(max_row=40)]
    header_idx, col_map = _find_header(rows)
    if header_idx is None:
        return {"error": "未识别到表头（需要包含 星期一~星期日 列）", "courses": [], "warnings": []}
    courses, warnings = _rows_to_courses(rows, header_idx, col_map)
    return {"sheet": ws.title, "header_row": header_idx + 1, "courses": courses, "warnings": warnings}


def _parse_xls(data: bytes) -> dict:
    """老版 Excel 97-2003 二进制格式（.xls），用 xlrd 解析。"""
    import xlrd

    book = xlrd.open_workbook(file_contents=data)
    sheet = book.sheet_by_index(0)
    rows = [
        [_normalize(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
        for r in range(min(sheet.nrows, 40))
    ]
    header_idx, col_map = _find_header(rows)
    if header_idx is None:
        return {"error": "未识别到表头（需要包含 星期一~星期日 列）", "courses": [], "warnings": []}
    courses, warnings = _rows_to_courses(rows, header_idx, col_map)
    return {"sheet": sheet.name, "header_row": header_idx + 1, "courses": courses, "warnings": warnings}


def _parse_csv(data: bytes) -> dict:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("gb18030", errors="replace")  # 国内 Excel 另存的 CSV 常为 GBK
    reader = csv.reader(io.StringIO(text))
    rows = [row for row in reader][:40]
    header_idx, col_map = _find_header(rows)
    if header_idx is None:
        return {"error": "未识别到表头（需要包含 星期一~星期日 列）", "courses": [], "warnings": []}
    courses, warnings = _rows_to_courses(rows, header_idx, col_map)
    return {"sheet": "CSV", "header_row": header_idx + 1, "courses": courses, "warnings": warnings}
